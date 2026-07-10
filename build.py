# 歴代役員経歴名簿.xlsx から index.html にデータを注入する
# 職場名は組合員名簿accdb（部局所属名）から取得（現組合員のみ）
# 使い方: python build.py  → git commit & push で公開反映
import csv
import json
import re
import subprocess
import tempfile
import unicodedata
from pathlib import Path

import openpyxl

XLSX = Path.home() / 'OneDrive/デスクトップ/歴代役員経歴名簿.xlsx'
ACCDB = Path.home() / 'OneDrive/デスクトップ/組合員名簿2026.7.1-ULTLA-KITAMATAPC-NECNOTE-KITAMATAPC-NECNOTE-NECNOTE-NECNOTE.accdb'
PS32 = r'C:\Windows\SysWOW64\WindowsPowerShell\v1.0\powershell.exe'
HTML = Path(__file__).parent / 'index.html'

# 異体字を標準形に潰す（照合キー用）
FOLD = str.maketrans(
    '﨑嵜髙濵濱邊邉齋齊斉瀨栁德眞冨桒藪籔淺澤薗嶌嶋曻惠廣國瀧榮壽靜萬與來兒龜檜條曾繩𠮷',
    '崎崎高浜浜辺辺斎斎斎瀬柳徳真富桑薮薮浅沢園島島昇恵広国滝栄寿静万与来児亀桧条曽縄吉')


def load_members():
    """accdbから組合員リスト [(氏名, ｼﾒｲ, 支部, 職場)] を作る。読めなければ空。"""
    if not ACCDB.exists():
        print('注意: 組合員名簿accdbが見つからない。職場名なしで続行')
        return []
    tmp = Path(tempfile.mkdtemp())
    out_csv = tmp / 'members.csv'
    script = tmp / 'export.ps1'
    ps = f'''$conn = New-Object System.Data.OleDb.OleDbConnection("Provider=Microsoft.ACE.OLEDB.16.0;Data Source={ACCDB}")
$conn.Open()
$cmd = $conn.CreateCommand()
$cmd.CommandText = "SELECT 氏名, ｼﾒｲ, 支部名称, 部局所属名 FROM [Sheet1]"
$r = $cmd.ExecuteReader()
$sb = New-Object System.Text.StringBuilder
while($r.Read()){{
  $vals = 0..3 | ForEach-Object {{ ('"' + ("$($r.GetValue($_))" -replace '"','') + '"') }}
  [void]$sb.AppendLine($vals -join ',')
}}
$r.Close(); $conn.Close()
[System.IO.File]::WriteAllText("{out_csv}", $sb.ToString(), (New-Object System.Text.UTF8Encoding $true))
'''
    script.write_bytes(b'\xef\xbb\xbf' + ps.encode('utf-8'))
    r = subprocess.run([PS32, '-NoProfile', '-ExecutionPolicy', 'Bypass', '-File', str(script)],
                       capture_output=True)
    if r.returncode != 0 or not out_csv.exists():
        print('注意: accdb読み取り失敗。職場名なしで続行')
        return []
    rows = []
    with open(out_csv, encoding='utf-8-sig') as f:
        for row in csv.reader(f):
            if row and row[0].strip():
                rows.append((row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()))
    return rows


def to_hira(kana):
    """半角ｶﾅ→ひらがな（濁点結合込み）"""
    s = unicodedata.normalize('NFKC', kana).replace(' ', '').replace('　', '')
    s = unicodedata.normalize('NFC', s)
    return ''.join(chr(ord(c) - 0x60) if 'ァ' <= c <= 'ヶ' else c for c in s)


def name_key(name):
    return name.replace('　', '').replace(' ', '').translate(FOLD)


member_rows = load_members()
members = {}
_tmp = {}
for _n, _k, _b, _w in member_rows:
    _tmp.setdefault(name_key(_n), set()).add(_w)
# 同姓同名で職場が割れる場合は使わない
members = {k: next(iter(v)) for k, v in _tmp.items() if len(v) == 1}

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.worksheets[0]
r1 = [c.value for c in ws[1]]
r2 = [c.value for c in ws[2]]

# 1行目のグループ名（中央執行部・各支部・青年部・女性部）を前方埋め
grp = ''
groups = []
for i in range(len(r2)):
    if r1[i]:
        grp = str(r1[i]).strip()
    groups.append(grp)

# 役職列 = E列(4)〜女性部常任幹事まで。集計列(中央四役〜)以降は除外
roles = []
for i in range(4, len(r2)):
    name = str(r2[i]).replace('\n', '').replace('　', '').strip() if r2[i] else ''
    if name in ('中央四役', '支部四役', '青年部四役', '女性部四役', '通算総数'):
        break
    roles.append((i, groups[i], name))
LAST = roles[-1][0]
TOTAL, MEMO = LAST + 9, LAST + 11  # 集計8列の後: 通算総数, (2026現役), 備考

people = []
matched = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[1]:
        continue
    rl = []
    for i, g, rn in roles:
        v = row[i]
        if v not in (None, ''):
            try:
                t = int(v)
            except (TypeError, ValueError):
                t = str(v)
            rl.append({'g': g, 'r': rn, 't': t})
    s, m = str(row[1]).strip(), str(row[2] or '').strip()
    w = members.get((s + m).translate(FOLD), '')
    if w:
        matched += 1
    people.append({
        'k': str(row[0] or ''), 's': s, 'm': m,
        'b': str(row[3] or ''), 'w': w, 'roles': rl,
        'total': int(row[TOTAL] or 0),
        'memo': str(row[MEMO] or ''),
    })

# 役員歴のない現組合員も検索に出す
xlsx_keys = {name_key(p['s'] + p['m']) for p in people}
added = 0
for _n, _kana, _b, _w in member_rows:
    if name_key(_n) in xlsx_keys:
        continue
    parts = _n.replace('　', ' ').split()
    s = parts[0]
    m = ' '.join(parts[1:]) if len(parts) > 1 else ''
    people.append({
        'k': to_hira(_kana), 's': s, 'm': m,
        'b': _b, 'w': _w, 'roles': [],
        'total': 0, 'memo': '',
    })
    added += 1
people.sort(key=lambda p: (p['k'] or 'ん'*10, p['s'], p['m']))

data = json.dumps(people, ensure_ascii=False, separators=(',', ':'))
html = HTML.read_text(encoding='utf-8')
new = re.sub(
    r'(<script id="data" type="application/json">).*?(</script>)',
    lambda m: m.group(1) + data + m.group(2),
    html, count=1, flags=re.S,
)
HTML.write_text(new, encoding='utf-8')
print(f'{len(people)}名を index.html に注入した（職場名あり {matched}名、役員歴なし組合員 {added}名追加）')
